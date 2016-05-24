# -*- coding: utf-8 -*- 
from django.db.models import Q
import xlrd,datetime
from django.core.exceptions import MultipleObjectsReturned
from unidecode import unidecode
import tempfile
import zipfile
import re
import random
#from pytz import timezone
from time import strftime, strptime
from django.utils import timezone
import pytz
import os
from exceptions import AttributeError
from django.forms.fields import DateField
from django.utils.timezone import localtime
from django.http.response import HttpResponse
from itertools import chain
from django.utils.safestring import mark_safe
SETTINGS_DIR = os.path.dirname(__file__)
MEDIA_ROOT = os.path.join(SETTINGS_DIR, 'media')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'LearnDriving.settings')
from rnoc.forms import D4_DATETIME_FORMAT, D4_DATE_ONLY_FORMAT
from rnoc.models import Tram, Mll, DoiTac, SuCo,\
    CaTruc, UserProfile, TrangThai, DuAn, ThaoTacLienQuan, ThietBi,\
    EditHistory, Lenh, FaultLibrary, NguyenNhan, Tinh, BSCRNC,\
    SiteType, BCNOSS, BTSType, UPE, QuanHuyen


DATE_FORMAT_FOR_BCN = '%d/%m/%Y'
TIME_FORMAT_FOR_BCN =  '%H:%M'
DATETIME_FORMAT_FOR_BCN = '%d/%m/%Y %H:%M:%S'
def unique_list(seq):
    seen = set()
    seen_add = seen.add
    return [x for x in seq if not (x in seen or seen_add(x))]

def read_line(path,split_item):
    f =  open(path, "r")
    content = f.read().decode('utf-8')
    content = content.split(split_item)
    f.close()
    return content


def read_file_from_disk (path):
    f =  open(path, "rb") 
    a = f.read().decode('utf-8')
    f.close()
    return a


def save_file_to_disk(path,content, is_over_write):
    if is_over_write:
        with open(path, "wb") as f:
            f.write(content.encode('utf-8'))
    else:
        with open(path, "ab") as f:
            f.write(content.encode('utf-8'))
import pytz
'''
def convert_awaredate__time_to_local(d):
    est=pytz.timezone('US/Eastern')
    d.astimezone(est)
    return d
'''
def awaredate_time_to_local(d):
    d = localtime(d)
    return d
def local_a_naitive(d,timezone = 'Asia/Bangkok'):
    eastern = pytz.timezone(timezone)
    loc_dt = eastern.localize(d)
    return loc_dt
def read_excel_cell(worksheet,row_index,curr_col):
    print  'curr_col row_index ',curr_col, row_index
    cell_value = worksheet.cell_value(row_index, curr_col)
    print 'cell_value',cell_value
    return cell_value

class Excel_2_3g(object):
    check_type_for_BCN= False
    allow_create_one_instance_if_not_exit = True
    fields_allow_empty_use_function =[]# nhung cai field ma excel = rong van dung fucntion de gan gia tri cho field, vi du nhu field namekhong dau
    is_import_from_exported_file = False
    #added_foreinkey_types = set() # cai nay dung de tinh so luong du an, hoac thietbi, duoc add, neu nhieu qua thi stop
    added_foreinkey_types = 0 # cai nay dung de tinh so luong du an, hoac thietbi, duoc add, neu nhieu qua thi stop
    max_length_added_foreinkey_types = 500
    backwards_sequence =[]
    many2manyFields = []
    update_or_create_main_item = ''#Site_ID_3G
    worksheet_name = u'Sheet1'
    begin_row=0
    manual_mapping_dict = {}
    mapping_function_to_value_dict = {}
    auto_map = False
    model = Tram
    created_number =0
    update_number = 0
    just_create_map_field = False
    
    def __init__(self,workbook=None,import_ghi_chu=None):
        
        self.import_ghi_chu = import_ghi_chu   
        self.workbook = workbook
        self.worksheet = self.workbook.sheet_by_name(self.worksheet_name)
        
        self.num_rows = self.worksheet .nrows - 1
        self.num_cols = self.worksheet.ncols - 1
        self.excel_dict = self.read_excel_dict() # excel_dict la ten cua cac cot lay trong file excel ra
        #print 'excel_dict',self.excel_dict
        #return false
        self.base_fields = {}
        self.missing_fields =[]
        self.model_fieldnames = [f.name for f in self.model._meta.fields if f.name!='id' ]
        if self.many2manyFields: # self.ModelClass._meta.many_to_many
                for x in self.many2manyFields:
                    if x not in self.model_fieldnames:
                        self.model_fieldnames.append(x)
        if self.is_import_from_exported_file:
            for f in  self.model._meta.fields:
                if (f.verbose_name in self.excel_dict)  :
                    if f.name =='id':
                        continue
                    else:
                        #print '@@@f.name ',f.name 
                        self.base_fields[f.name] = self.excel_dict.get(f.verbose_name)
                elif f.name.replace("_",' ') in self.excel_dict:
                    self.base_fields[f.name] = self.excel_dict.get(f.name.replace("_",' ') )
                else:
                    self.missing_fields.append(f.name)
        else:
            for fname in self.model_fieldnames:
                fname_lower = fname.lower()
                if self.auto_map and (fname_lower in self.excel_dict):
                    self.base_fields[fname] = self.excel_dict[fname_lower]
                else: # 1 so attribute khong nam trong file excel
                    if fname in self.manual_mapping_dict: #manual_mapping_dict la manual , do minh tu tao anh xa fieldname voi ten cot cua file excel
                        fieldname_compare_with_fn_excel = self.manual_mapping_dict[fname]
                        if isinstance(fieldname_compare_with_fn_excel, int):
                            self.base_fields.update({fname:fieldname_compare_with_fn_excel})
                        else:
                            fieldname_in_excel =  unidecode(fieldname_compare_with_fn_excel).lower().replace(' ','_') # file name format
                            if fieldname_in_excel in self.excel_dict:
                                self.base_fields[fname]= self.excel_dict[fieldname_in_excel] #= so thu tu cua column chua field do, vi du 5
                            else: # thieu cot nay hoac da bi doi ten                        
                                raise ValueError('trong file excel thieu cot %s '%fieldname_in_excel)
                    else:
                        self.missing_fields.append(fname)
        for fname in self.fields_allow_empty_use_function:
            if fname in self.model_fieldnames and fname not in self.base_fields:
                self.base_fields[fname] = None
        #print '@@@@@@@@@@@@@base_fields',self.base_fields
        
        self.convert_basefield_to_list_of_tuple()
        
        if self.just_create_map_field:#for test
            return None
        
        self.loop_through_row_and_insertdb()
    def read_excel_dict(self): #{verbose_name_in_excel: col_index,verbose_name_in_excel2: col_index2}
        dict_attrName_columnNumber_excel_not_underscore = {}
        row_index = self.begin_row
        curr_col = 0
        while curr_col <= self.num_cols:
            atrrname = read_excel_cell(self.worksheet, row_index,curr_col)
            #atrrname la field name hay la collumn name
            if not self.is_import_from_exported_file:
                #print '@@@@@@@@@@@atrrname',curr_col,atrrname
                atrrname = unidecode(atrrname).lower().replace(" ","_")
            else:
                atrrname = atrrname.replace("_"," ")
                
            dict_attrName_columnNumber_excel_not_underscore[atrrname ]=   curr_col
            curr_col +=1
        return dict_attrName_columnNumber_excel_not_underscore
    def convert_basefield_to_list_of_tuple(self):
        #self.main_field_index_excel_column = self.base_fields.pop(self.update_or_create_main_item) #index of main fields
        self.main_dict = {}
        if isinstance(self.update_or_create_main_item, str):
            self.update_or_create_main_item = (self.update_or_create_main_item,)
            print 'wwwwwwwwwwwwwwwwself.update_or_create_main_item ',self.update_or_create_main_item 
        for i in self.update_or_create_main_item:
            self.main_dict.update({i:self.base_fields.pop(i)})
        
        
        
        if self.backwards_sequence:
            #can phai sap xep lai theo chuan [(),()] theo thu tu
            base_fields_lists = [x for x in self.base_fields.iterkeys()]
            for x in self.backwards_sequence:
                if x in base_fields_lists:
                    base_fields_lists.remove(x)
                    base_fields_lists.append(x)
            self.odering_base_columns_list_tuple = [(x, self.base_fields[x]) for x in base_fields_lists]
        else:
            self.odering_base_columns_list_tuple = [(k,v) for k, v in self.base_fields.iteritems()]
        #print 'self.odering_base_columns_list_tuple',self.odering_base_columns_list_tuple
    
    
    def loop_through_row_and_insertdb(self):
        print 'loop_through_row_and_insertdb'
        row_index = getattr(self,'row_index',self.begin_row) 
        self.tram_co_trong_3g_location_but_not_in_db3g = 0
        karg = {}
        #print 'self.num_rows',self.num_rows
        self.has_created_tram_instance = False
        while row_index < self.num_rows:
            row_index += 1
            print 'row_index',row_index
            if self.check_type_for_BCN:
                n_nhan  = read_excel_cell(self.worksheet,row_index, self.excel_dict['n.nhan'])
                thoi_gian_cb  = read_excel_cell(self.worksheet,row_index, self.excel_dict['thoi_gian_cb'])
                comment_vnp =  read_excel_cell(self.worksheet,row_index, self.excel_dict['vnp-ghi_chu'])
                if (thoi_gian_cb!='' and int(thoi_gian_cb) < 10) or n_nhan !='SITE_OOS' or comment_vnp =='':
                    continue
                loai_ne = read_excel_cell(self.worksheet,row_index, self.excel_dict['loai_ne'])
                nha_cc = read_excel_cell(self.worksheet,row_index, self.excel_dict['nha_cc'])
                if loai_ne=='BSC':
                    continue
                
                if (loai_ne=='CELL' and nha_cc=='Ericsson'):
                    self.type_excel = '3G'
                elif(loai_ne=='NODEB' and nha_cc=='Nokia'):
                    self.type_excel = 'NSM'
                elif(loai_ne=='NODEB' and nha_cc=='Alcatel'):
                    self.type_excel = 'ALU'
                elif(loai_ne=='BTS' and nha_cc=='Ericsson'):
                    self.type_excel = 'SRN'
                else:
                    self.type_excel = '2G'
                 
            for main_field in self.main_dict:
                value = read_excel_cell(self.worksheet,row_index, self.main_dict[main_field])
                to_value_function = self.get_function(main_field) # function for main field
                if to_value_function:
                    value = to_value_function(value)
                    karg.update({main_field:value})
            filters = self.model.objects.filter(**karg)
            if filters: # co db_row nay roi, update thoi
                #print '*update'
                self.created_or_update = 0
                for self.obj in filters:# loop va gan gia tri vao self.obj
                    self.setattr_field_for_obj(row_index)
                    self.update_number +=1
                    #print '@ so luong instance duoc update,',self.update_number
            else: #tao moi
                #print '*tao moi'
                if self.allow_create_one_instance_if_not_exit:
                    self.created_or_update = 1   
                    self.obj = self.model(**karg)
                    self.setattr_field_for_obj(row_index)
                    self.created_number +=1
                    #print '@ so luong instance duoc tao moi(new),',self.created_number
                else:
                    self.tram_co_trong_3g_location_but_not_in_db3g +=1
                    #print '@ so luong instance tram_co_trong_3g_location_but_not_in_db3g,',self.tram_co_trong_3g_location_but_not_in_db3g
                    continue
        self.thong_bao =  u'''số dòng được đọc %s,
        số dòng được cập nhập %s, 
        Số dòng được lưu vào database %s'''%(row_index,self.update_number,self.created_number )
        #print '***Tong Ket '
        #print '@ so luong instance duoc update,',self.update_number
        #print '@ so luong instance duoc tao moi(new),',self.created_number
        #print '@ so luong instance tram_co_trong_3g_location_but_not_in_db3g,',self.tram_co_trong_3g_location_but_not_in_db3g
    def setattr_field_for_obj(self,row_index):
        for field_tuple in self.odering_base_columns_list_tuple:
            field = field_tuple[0]
            column_number = field_tuple[1]
            #print 'field',field
            if column_number !=None:
                value =  read_excel_cell(self.worksheet, row_index,column_number)
            else:
                value = None
            if value=='' or value =="null"or value ==u'—':
                value = None
            if  value !=None or field in self.fields_allow_empty_use_function:
                to_value_function = self.get_function(field)
                if to_value_function:
                    value = to_value_function(value)
                    if value==None:#to_value_function chu dong tra ve None neu khong muon luu field
                        continue
                setattr(self.obj, field, value) # save
           
        self.obj.save()        
    def get_function(self,field):
        if field in self.mapping_function_to_value_dict:
            func_name = self.mapping_function_to_value_dict[field]
            to_value_function = getattr(self, func_name)
            return to_value_function
        else:
            try:
                method_of_field_name = 'value_for_'+field
                to_value_function = getattr(self, method_of_field_name) #
                return to_value_function
            except: # Ko co ham nao thay doi gia tri value
                return None
            
    def value_for_tinh(self,value):
        print '@@@@@@@@@@@@@2tinh le',value
        if value:
            try:
                instance = Tinh.objects.get(ma_tinh = value)
            except:
                return None
                self.added_foreinkey_types +=1
                if self.added_foreinkey_types > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance = Tinh(ma_tinh = value)
                instance.save()
            self.obj
            return instance
        else:
            return None        
    def value_for_import_ghi_chu(self,value,prefix='3G: ',insert_index=0):
        import_ghi_chu_old = getattr(self.obj, 'import_ghi_chu',None)
        #print '@@@@@@import_ghi_chu_old',import_ghi_chu_old
        import_ghi_chu_old_s=[]
        if import_ghi_chu_old:
            import_ghi_chu_old_s = import_ghi_chu_old.split('\n')
        now_string = datetime.datetime.now().strftime(D4_DATETIME_FORMAT)
        string = prefix + u'import từ file ' + self.import_ghi_chu + u' vào lúc ' + now_string
        #print '@string',string
        if prefix=='':
            raise ValueError('prefix must define in agrument')
        else:
            try: 
                x = import_ghi_chu_old_s[insert_index]
                if prefix in x[:4]:
                    #print 'chi chong'
                    import_ghi_chu_old_s[insert_index] = string
                else:
                    import_ghi_chu_old_s.insert(insert_index,string)# gia su 4G phai nam o index 2 nhung no dang nam o 1, ma dang import_ghi_chu 2G
                    
            except IndexError:
                import_ghi_chu_old_s.insert(insert_index,string)
                
        
        #import_ghi_chu_old_s.insert(insert_index,string)
        return_value = '\n'.join(import_ghi_chu_old_s)
        #print '@@@return_value import_ghi_chu_old_s',return_value
        return return_value
    def value_for_Site_type(self,value):
        return SiteType.objects.get_or_create(Name = u'Site thường')[0]
    def value_for_dateField(self,cell_value):
        try:
            date = datetime.datetime(1899, 12, 30)
            get_ = datetime.timedelta(int(cell_value)) # delte du lieu datetime
            get_col2 = str(date + get_)[:10] # convert date to string theo dang nao do
            value = get_col2 # moi them vo
            return value
        except:
            return None
    def value_for_RNC(self,value):
        if value:
            try:
                instance = Tram.objects.get(Site_Name_1 = value)
            except Tram.DoesNotExist:
                if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance = Tram(Site_Name_1 = value,Site_type = SiteType.objects.get(Name = u'Site 0 (RNC,BSC)'),ngay_gio_tao = timezone.now(),nguoi_tao = User.objects.get(username = u'rnoc2'))
                instance.save()
                
                
            try:
                instance = BSCRNC.objects.get(Name = value)
            except BSCRNC.DoesNotExist:
                if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance = BSCRNC(Name = value, ngay_gio_tao = timezone.now(),nguoi_tao = User.objects.get(username = 'rnoc2'))
                instance.save()
            return instance
        else:
            return None
    def value_for_Name(self,value):
        if value:
            return value.rstrip().lstrip()
        else:
            return ''
    def value_for_common_datefield(self,cell_value):
        try:
            date = datetime.datetime(1899, 12, 30)
            get_ = datetime.timedelta(int(cell_value)) # delta du lieu datetime
            value = date + get_
            return value
        except:
            return None
    def value_for_common_datefield_exported_type(self,cell_value):
        cell_value = re.sub("$'", "", cell_value)
        d = datetime.datetime.strptime(cell_value, '%d/%m/%Y')
        return d
    def value_for_Cabinet(self,cell_value,name_ThietBi_attr= 'Cabinet',bts_type = None):
        try:
            thietbi = ThietBi.objects.get(Name=cell_value,bts_type = bts_type)
        except:
            thietbi = ThietBi(Name=cell_value)
            self.added_foreinkey_types +=1
            if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
            thietbi.is_duoc_tao_truoc = True
            thietbi.nguoi_tao = User.objects.get(username="rnoc2")
            thietbi.bts_type = bts_type
            thietbi.ly_do_sua = u'Được tạo ra từ import database Trạm'
            thietbi.save()
        setattr(self.obj,name_ThietBi_attr,thietbi)
        return None
    def value_for_common_VLAN_ID (self,cell_value):
        value = int(cell_value)
        return value   
    
    def value_for_nguoi_tao (self,cell_value):
        #print '@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ value_for_nguoi_tao',self.created_or_update 
        if self.created_or_update ==1:
            if cell_value==None:
                user = User.objects.get (username = 'rnoc2')
                return user
            else:
                user = User.objects.get (username = cell_value)
                return user
        else:
            return None
    def value_for_nguoi_sua_cuoi_cung (self,cell_value):
        user = User.objects.get (username = cell_value)
        return user

    def value_for_ngay_gio_tao(self,cell_value):
        if self.created_or_update ==1:
            if cell_value:
                cell_value = re.sub("^'", "", cell_value)
                cell_value = re.sub('\s$', '', cell_value)
                d = datetime.datetime.strptime(str(cell_value), D4_DATETIME_FORMAT)
                '''eastern = pytz.timezone('Asia/Bangkok')
                loc_dt = eastern.localize(d)
                '''
                #loc_dt = timezone.localtime(d)
                return d
            else:
                #now = datetime.datetime.now()
                now = timezone.now()
                '''
                now_str = now.strftime(D4_DATETIME_FORMAT)
                now = datetime.datetime.strptime(now_str, D4_DATETIME_FORMAT)
                #print '@@@@@@@@type cua now',type(now)
                '''
                return now
        else:
            return None
    def value_for_ngay_gio_sua(self,cell_value):
        cell_value = re.sub("^'", "", cell_value)
        cell_value = re.sub(' $', '', cell_value)
        d = datetime.datetime.strptime(str(cell_value), D4_DATETIME_FORMAT)
        return d
     
    def value_for_excel_export_boolean(self,cell_value):#boolean not null allowed
        #print '@@@@@@@@@cell_value',cell_value
        if cell_value ==u'✘' or cell_value==None:
            return False
        else:
            return True
    def value_for_Name_khong_dau(self,cell_value):
        field_co_dau = self.obj.Name
        if field_co_dau==None:
            return None
        return unidecode(field_co_dau)
class ExcelChung (Excel_2_3g):
    #backwards_sequence =['Site_ID_2G',]#de lay gia tri nha_san_xuat_2G truoc
    is_import_from_exported_file = True
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Database 2G'
    mapping_function_to_value_dict ={'Ngay_Phat_Song_2G':'value_for_common_datefield_exported_type','Ngay_Phat_Song_3G':'value_for_common_datefield_exported_type'}
    manual_mapping_dict = {}

    def value_for_nha_san_xuat_2G(self,cell_value):
        thietbi = ThietBi.objects.get_or_create(Name=cell_value)[0]
        self.added_foreinkey_types.add(thietbi)#set().add
        l = len(self.added_foreinkey_types)
        #print "cabin**",l
        if l >self.max_length_added_foreinkey_types:
            raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
        self.obj.nha_san_xuat_2G=thietbi
        #self.obj.save()
        return None
    
    
    



    
class ExcelImportTrangThai (Excel_2_3g):
    is_import_from_exported_file = True
    fields_allow_empty_use_function = ['color_code','is_cap_nhap_gio_tot','nguoi_tao','ngay_gio_tao','is_duoc_tao_truoc','Name_khong_dau']
    #backwards_sequence =['Name_khong_dau']#de lay gia tri nha_san_xuat_2G truoc
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Name'
    worksheet_name = u'Sheet 1'
    mapping_function_to_value_dict ={'is_cap_nhap_gio_tot':'value_for_excel_export_boolean','is_duoc_tao_truoc':'value_for_excel_export_boolean'}
    manual_mapping_dict = {}
    model = TrangThai
    
    def value_for_color_code(self,cell_value):#boolean
        if cell_value:
            return cell_value
        else:
            return "#%06x" % random.randint(0, 0xFFFFFF)
    '''
    def value_for_is_cap_nhap_gio_tot(self,cell_value):#boolean not null allowed
        #if cell_value ==u'✔':
            #return True
        if cell_value ==u'✘' or cell_value==None:
            return False
        else:
            return True
    
    def value_for_is_duoc_tao_truoc(self,cell_value):
        if self.obj.Name  ==u"Raise sự kiện":
            return True
        else:
            return False
    '''
class ExcelImportSuCo(ExcelImportTrangThai):
    model = SuCo
'''
class ExcelImportSuCo_old(Excel_2_3g):
    is_import_from_exported_file = True
    fields_allow_empty_use_function = ['nguoi_tao','ngay_gio_tao','Name_khong_dau']
    backwards_sequence =[]
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Name'
    worksheet_name = u'Sheet 1'
    mapping_function_to_value_dict ={}
    manual_mapping_dict = {}
    model = SuCo
'''
class ExcelImportNguyenNhan(ExcelImportTrangThai):
    model = NguyenNhan
    def value_for_color_code(self,cell_value):#boolean
        if cell_value:
            return cell_value
        else:
            #return "#%06x" % random.randint(0, 0xFFFFFF)
            return 'green'
class ExcelImportDoiTac (ExcelImportTrangThai):
    model = DoiTac
class ExcelImportDoiTac_ungcuu (Excel_2_3g):
    fields_allow_empty_use_function = ['Name_khong_dau','nguoi_tao','ngay_gio_tao']
    manual_mapping_dict = {'Name':u'HO_TEN','Don_vi' :u'TINH_TP1','So_dien_thoai':u'DIEN_THOAI','Thong_tin_khac':u'TINH_TP','email':u'EMAIL'}
    model = DoiTac
    worksheet_name = u'Sheet1'
    update_or_create_main_item = 'Name'
    def value_for_So_dien_thoai(self,value):
        if value:
            value = str(value)
            
            rs = re.subn('\.0$', '', value, 1)
            if rs[1]:
                value = rs[0]
                if value[0] !='0':
                    value = '0' + value
                return value
            else:
                return value
        else:
            return None
    def value_for_Thong_tin_khac(self,value):
        if value:
            try:
                tinh = Tinh.objects.get(ma_tinh=value)
                value = value + u', ' + tinh.dia_ban
                return value
            except Tinh.DoesNotExist:
                return value
        else:
            return None
class ImportTinh(Excel_2_3g):
    fields_allow_empty_use_function = ['Name_khong_dau']
    manual_mapping_dict = {'Name':u'Tên tỉnh','ma_tinh' :u'TINH_TP',}
    model = Tinh
    worksheet_name = u'Sheet1'
    update_or_create_main_item = 'Name'
DATETIME_FORMAT_BCN = '%d/%m/%Y %H:%M:%S'
class ImportBCN2G(Excel_2_3g):
    #row_index = 1276
    check_type_for_BCN= True
    type_excel = ''
    backwards_sequence = ['BTS_thiet_bi']
    thietbi_add_count = 0
    begin_row=12
    fields_allow_empty_use_function = ['BSC_or_RNC']
    manual_mapping_dict = {'object':u'Tên NE','gio_mat' :u'Thời gian sự cố','gio_tot':u'Thời gian CLR',\
                           'code_loi' :u'Loại sự cố','vnp_comment':u'VNP-Ghi chú',\
                           'gio_canh_bao_ac' :u'Thời gian cảnh báo AC','tong_thoi_gian':u'Thời gian CB','BTS_thiet_bi':u'Nhà CC' ,'BTS_Type':u'Loại NE',
                           }
    model = BCNOSS
    #worksheet_name = u'Sheet1'
    update_or_create_main_item = (u'object','gio_mat')
    def value_for_tong_thoi_gian(self,value):
        if value:
            return int(value)
        else:
            return None
    
    def value_for_BTS_thiet_bi(self,value):
        try:
            thietbi = ThietBi.objects.get(Name=value,bts_type = self.obj.BTS_Type )
        except:
            thietbi = ThietBi(Name=value,bts_type = self.obj.BTS_Type )
            self.thietbi_add_count +=1
            if self.thietbi_add_count >10:
                raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
            thietbi.is_duoc_tao_truoc = True
            thietbi.nguoi_tao = User.objects.get(username="rnoc2")
            thietbi.save()
        
        if self.type_excel == 'ALU' and self.has_created_tram_instance:
            self.tram_instance.Cabinet = thietbi
            self.tram_instance.save()
            self.has_created_tram_instance = False
        return thietbi
        #return ThietBi.objects.get(Name = value)
    
    def value_for_BTS_Type(self,value):
        if self.type_excel == '3G' or self.type_excel == 'ALU'  or self.type_excel == 'NSM':
            value = BTSType.objects.get(Name = '3G')
            return value
        if value == 'BTS':
            value = BTSType.objects.get(Name = '2G')
            return value
    
    def value_for_object(self,value):#tra ve bsc rnc luon
        if self.type_excel == 'ALU':
            object_name = value
            site_name1_for_look  =re.sub('^3G_','',object_name)
            print 'site_name1_for_look',site_name1_for_look
            try:
                tram_alu = Tram.objects.filter(Site_Name_1 = site_name1_for_look)[0]
                bsc_or_rnc = tram_alu.RNC
            except: #IndexError,DoesNotExist:
                user = User.objects.get(username="rnoc2")
                now = timezone.now()
                '''
                try:
                    bsc_or_rnc = BSCRNC.objects.get(Name = "ALURNC")
                except:
                    bsc_or_rnc = BSCRNC(Name = "ALURNC")
                    bsc_or_rnc.nguoi_tao = user
                    bsc_or_rnc.ngay_gio_tao = now
                    bsc_or_rnc.ly_do_sua = u'Được tạo ra từ import báo cáo ngày %s'%now
                    bsc_or_rnc.save()
                '''
                bsc_or_rnc = None
                self.tram_instance = Tram(Site_Name_1 = site_name1_for_look,Site_ID_3G = 'ALU_'+ object_name,RNC=bsc_or_rnc,\
                     nguoi_tao=user,ngay_gio_tao=now,ly_do_sua=u'Được tạo ra từ import báo cáo ngày %s'%now.strftime(D4_DATETIME_FORMAT),Site_type = SiteType.objects.get(Name = 'Site thường'))
                self.has_created_tram_instance = True
                #instance.save()
            self.BSC_or_RNC = bsc_or_rnc
            #self.BTS_Type = BTSType.objects.get(Name = '3G')
            return object_name
        elif self.type_excel == 'NSM':
            pattern = u'^(.*?) (.*?)$'#HCRNC23 3G_BCH065K7_HCM
            m = re.search(pattern, value)
            object_name = m.group(2) 
            bsc_rnc_name=  m.group(1).replace('-','')
            #print instance
            #self.BSC_or_RNC = instance
            #return object_name
            #self.BTS_Type = BTSType.objects.get(Name = '3G')
        elif self.type_excel == '3G':
            pattern = u'^(.*?) (.*?)\d(_.*?)$'#HCRNC23 3G_BCH065K7_HCM
            m = re.search(pattern, value)
            object_name = m.group(2) +  m.group(3)
            bsc_rnc_name=  m.group(1)
            #self.BTS_Type = BTSType.objects.get(Name = '3G')
        else:#2G
            #pattern_lists = [u'^(.*?)\(.*?[: ](.*?)\)'] #BSC_753M_TVH(BTS_35:TCA020A_TVH) ,BSC_405H_LAN(XVT-TEST-BTS3900_LAN)
            if self.type_excel =='SRN':#HCBS249_HCM 2G_HMO011E_HCM
                #,u'^(.*?)(.*?)$'
                pattern = u'^(.*?) (.*?)$'
                kq = re.findall(pattern, value)
            elif self.type_excel =='2G':
                pattern_lists =[u'^(.*?)\(.*?[: ](.*?)\)',u'^(.*?)\((.*?)\)$'] #u'^(.*?)\((.*?)\)$'  cho truong hop BSC_401H_KGG(LD-RACH-GIA_KGG)
                for pattern in pattern_lists:
                    kq = re.findall(pattern, value)
                    if kq:
                        break
            bsc_rnc_name = kq[0][0]
            object_name =  kq[0][1]
            #self.BTS_Type = BTSType.objects.get(Name = '2G')
        try:
            #bsc_instance = BSCRNC.objects.get(Name__icontains =bsc_rnc_name)[0]
            bsc_instance = BSCRNC.objects.filter(Name__icontains = bsc_rnc_name)[0]
        except:
            if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
            bsc_instance = BSCRNC(Name = bsc_rnc_name, ngay_gio_tao = timezone.now(),nguoi_tao = User.objects.get(username = 'rnoc2')\
                                  ,ly_do_sua = u'được tạo ra khi import bao cao ngay %s'%self.type_excel)
            bsc_instance.save()
        self.BSC_or_RNC = bsc_instance
        return object_name
            
            
    def value_for_vnp_comment(self,value):
        #if value =='':
            #return '__skip__'
        m = re.search(u'^(.*?)- \d -', value)
        value  = m.group(1)
        return value
    def value_for_BSC_or_RNC(self,value):
        return self.BSC_or_RNC
    '''
    def value_for_BSC_or_RNC(self,value):
        if value:
            m = re.search(u'^(.*?)\(', value)
            value  = m.group(1)
            #print '@@@@@@@@@@@@@@@',m.group(1)
            #print '######rnc name',value
            try:
                instance = BSCRNC.objects.get(Name = value)
            except BSCRNC.DoesNotExist:
                if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance = BSCRNC(Name = value, ngay_gio_tao = timezone.now(),nguoi_tao = User.objects.get(username = 'rnoc2'))
                instance.save()
            return instance
        else:
            return None
    '''
    def value_for_gio_canh_bao_ac(self,value):
        if value:
            #print 'gio_canh bao ac_@@@',value,type(value)
            d = datetime.datetime.strptime(value, DATETIME_FORMAT_BCN)
            return d
        else:
            return None
        
    def value_for_gio_mat(self,value):
        #print 'gio_mat_@@@',value,type(value)
        d = datetime.datetime.strptime(value, DATETIME_FORMAT_BCN)
        return d
    def value_for_gio_tot(self,value):
        if value:
            #print 'gio_tot_@@@',value,type(value)
            d = datetime.datetime.strptime(value, DATETIME_FORMAT_BCN)
            return d
        else:
            return None
class ImportBCN3G(ImportBCN2G):
    pass
class ImportBCN2G_SRAN(ImportBCN2G):
    pass
class ImportBCN3G_ALU(ImportBCN2G):
    pass
class ImportBCN3G_NSM(ImportBCN2G):
    pass    
class ImportTinh_diaban(Excel_2_3g):
    fields_allow_empty_use_function = []
    manual_mapping_dict = {'Name':u'Khu vực','dia_ban' :u'Địa bàn','ghi_chu':u'Trực UCTT'}
    model = Tinh
    worksheet_name = u'Sheet3'
    update_or_create_main_item = 'Name'    
class ImportRNC(Excel_2_3g):
    auto_map = True
    fields_allow_empty_use_function = ['nguoi_tao','ngay_gio_tao','import_ghi_chu']
    manual_mapping_dict = {'Name':u'RNCID'}
    model = BSCRNC
    worksheet_name = u'Thong ke NodeB-RNC'
    update_or_create_main_item = 'Name'
    def value_for_VI_TRI_RNC(self,value):
        if value:
            return Tinh.objects.get(ma_tinh = value)
        else:
            return None
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_3G,self).value_for_import_ghi_chu(value,'RNC: ',insert_index=0)
        return return_value
class Import_RNC_Tram(Excel_2_3g):
    fields_allow_empty_use_function = ['nguoi_tao','ngay_gio_tao','Site_type']
    manual_mapping_dict = {'Site_Name_1':u'Name','dia_chi_2G':u'DIA CHI','dia_chi_3G':u'DIA CHI'}
    model = Tram
    worksheet_name = u'Sheet 1'
    update_or_create_main_item = 'Site_Name_1'
    def value_for_ngay_gio_tao(self,value):
        return timezone.now()
    def value_for_Site_type(self,value):
        return SiteType.objects.get_or_create(Name = u'Site 0 (RNC,BSC)')[0]
'''
class ExcelImportDoiTac_old (Excel_2_3g):
    is_import_from_exported_file = True
    fields_allow_empty_use_function = ['Name_khong_dau','nguoi_tao','ngay_gio_tao']
    backwards_sequence =['Name_khong_dau',]#de lay gia tri nha_san_xuat_2G truoc
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Name'
    worksheet_name = u'Sheet 1'
    mapping_function_to_value_dict ={}
    manual_mapping_dict = {}
    model = DoiTac
    def value_for_Name_khong_dau(self,cell_value):
        field_co_dau = self.obj.Name
        if cell_value:
            return unidecode(cell_value)
        else:
            return unidecode(field_co_dau)
        
'''
class ExcelImportDuAn(ExcelImportTrangThai):
    model = DuAn
    def value_for_color_code(self,cell_value):#boolean
        if cell_value:
            return cell_value
        else:
            #return "#%06x" % random.randint(0, 0xFFFFFF)
            return 'organe'
'''      
class ExcelImportDuAn_old(Excel_2_3g):
    is_import_from_exported_file = True
    fields_allow_empty_use_function = ['nguoi_tao','ngay_gio_tao','Name_khong_dau']
    backwards_sequence =[]#de lay gia tri nha_san_xuat_2G truoc 
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Name'
    worksheet_name = u'Sheet 1'
    mapping_function_to_value_dict ={'is_duoc_tao_truoc':'value_for_excel_export_boolean'}#uu tien sau
    manual_mapping_dict = {}
    model = DuAn
'''
class ExcelImportThietBi(ExcelImportTrangThai):
    model = ThietBi
    def value_for_color_code(self,cell_value):#boolean
        if cell_value:
            return cell_value
        else:
            #return "#%06x" % random.randint(0, 0xFFFFFF)
            return 'purple'
class ExcelImportFaultLibrary(ExcelImportTrangThai):
    model = FaultLibrary
class ExcelImportThaoTacLienQuan (ExcelImportTrangThai):
    model = ThaoTacLienQuan
class ExcelImportLenh (ExcelImportDuAn):
    fields_allow_empty_use_function = ['color_code','is_cap_nhap_gio_tot','nguoi_tao','ngay_gio_tao','is_duoc_tao_truoc','Name_khong_dau',]
    backwards_sequence= ['Name_khong_dau']
    update_or_create_main_item = 'command'
    model = Lenh
    def value_for_thiet_bi(self,value):
        if value:
            try:
                tb = ThietBi.objects.get(Name=value)
                return tb
            except ThietBi.DoesNotExist:
                return None
        else:
            return None
    '''    
    def value_for_Name(self,value):
        #print '@@@@@@@@i want see'
        if value==None:
            #print '@@@@@@@@i want see2'
            return ''
        else:
            return value
    '''
class Excel_3G(Excel_2_3g):
    #import_ghi_chu_text = 'Ericsson_Database_Ver_160'
    fields_allow_empty_use_function = ['nguoi_tao','ngay_gio_tao','active_3G','Site_type','import_ghi_chu']
    auto_map = True
    many2manyFields = ['du_an']
    just_create_map_field = False
    update_or_create_main_item = 'Site_ID_3G'
    worksheet_name = u'Ericsson 3G'
    backwards_sequence =['du_an','UPE']
    manual_mapping_dict = {'Project_Text':2,'du_an':2,'Cell_1_Site_remote':u'Cell 1 (carrier 1)', \
                    'Cell_2_Site_remote':u'Cell 2 (Carrier 1)', 'Cell_3_Site_remote':u'Cell 3 (Carrier 1)',\
                     'Cell_4_Site_remote':u'Cell 4 (Carrier 2)', 'Cell_5_Site_remote':u'Cell 5 (Carrier 2)', 'Cell_6_Site_remote':u'Cell 6 (Carrier 2)', \
                     'Cell_7_Site_remote':u'Cell 7 (remote/U900/3 carrier)', 'Cell_8_Site_remote':u'Cell 8 (remote/U900/3 carrier)', 'Cell_9_Site_remote':u'Cell 9 (remote/U900/3 carrier)',\
                     'Cell_K_U900_PSI':u'Cell K (U900 PSI)','tinh':u'Count Province'
                     }
    mapping_function_to_value_dict = {'Ngay_Phat_Song_2G':'value_for_dateField',\
                                      'Ngay_Phat_Song_3G':'value_for_dateField',\
                                      'IUB_VLAN_ID':'value_for_int_to_string','MUB_VLAN_ID':'value_for_int_to_string',\
                                      'Cell_1_Site_remote':u'value_error_but_equal_42', \
                    'Cell_2_Site_remote':u'value_error_but_equal_42', 'Cell_3_Site_remote':u'value_error_but_equal_42',\
                     'Cell_4_Site_remote':u'value_error_but_equal_42', 'Cell_5_Site_remote':u'value_error_but_equal_42', 'Cell_6_Site_remote':u'value_error_but_equal_42', \
                     'Cell_7_Site_remote':u'value_error_but_equal_42', 'Cell_8_Site_remote':u'value_error_but_equal_42', 'Cell_9_Site_remote':u'value_error_but_equal_42',\
                     'Cell_K_U900_PSI':u'value_error_but_equal_42'
                                    }
    def value_for_Cabinet(self,value):
        bts_type = BTSType.objects.get(Name = '3G')
        return super(Excel_3G,self).value_for_Cabinet(value,bts_type = bts_type)
    def value_for_UPE (self,value):
        if value:
            try:
                instance = UPE.objects.get(Name = value,tinh = self.obj.tinh )
            except:
                try:
                    instance = UPE(Name = value,tinh = self.obj.tinh)
                    instance.save()
                except:
                    instance = None
            return instance
        else:
            return None
        
    def value_for_import_ghi_chu(self,value,prefix='',insert_index=0):
        return_value = super(Excel_3G,self).value_for_import_ghi_chu(value,'3G: ',insert_index=0)
        return return_value
    def value_error_but_equal_42(self,value):
        if value ==42:
            return None
        else:
            return value
    def value_for_du_an(self,cell_value):
        
        if self.created_or_update == 1 :#create
            self.obj.save()
        try:
            du_an = DuAn.objects.get(Name=cell_value)
        except:
            du_an = DuAn(Name=cell_value)
            self.added_foreinkey_types +=1
            if self.added_foreinkey_types > self.max_length_added_foreinkey_types:
                raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
            du_an.type_2G_or_3G = '3G'
            du_an.is_duoc_tao_truoc = True
            du_an.nguoi_tao = User.objects.get(username="rnoc2")
            du_an.ngay_gio_tao = timezone.now()
            du_an.save()
        self.obj.du_an.add(du_an)
        return None
    
    
    def value_for_Site_ID_3G(self,cell_value):
        ex = re.subn('_U900$','',cell_value)
        if ex[1]:# co phat hien chuoi _U900 o trong
            self.is_U900_hay_U2100 = 'U900'
            same_3g_has_site_name1 = ex[0]
            try:# co tram U2100 tuong ung voi tram U900 nay
                samesite_instance = Tram.objects.get(Site_ID_3G = 'ERI_3G_' + same_3g_has_site_name1)
                #giasu u2100 luon duoc tao ra truoc,
                samesite_instance.is_co_U900_rieng = True
                samesite_instance.save()
                self.is_co_U2100_rieng = True
            except:
                self.is_co_U2100_rieng = False
        else:# tram U2100
            self.is_U900_hay_U2100 = 'U2100'
            try:# neu co tram U900 rieng
                samesite_instance = Tram.objects.get(Site_ID_3G = 'ERI_3G_' + cell_value + '_U900')
                samesite_instance.is_co_U2100_rieng = True
                self.is_co_U900_rieng = True
            except:# khogn co tram U900 nao ung voi tram U2100 nay
                self.is_co_U900_rieng = False
        value = 'ERI_3G_' + cell_value
        return value
    def value_for_Site_ID_2G(self,value):
        return 'SRN_2G_' + value
    def value_for_Site_Name_2(self,cell_value):
        value = cell_value.replace('3G_','')
        return value
    
    def value_for_int_to_string (self,cell_value):
        value = int(cell_value)
        return value
    def value_for_Site_Name_1 (self,value):
        
        
        if self.is_U900_hay_U2100 == 'U900':
            self.obj.is_co_U2100_rieng = self.is_co_U2100_rieng
        elif self.is_U900_hay_U2100 == 'U2100':
            self.obj.is_co_U900_rieng = self.is_co_U900_rieng
        value = value.replace("3G_","")
        return value
    def value_for_active_3G(self,value):
        if self.created_or_update == 1 :#create
            return True
        
    def value_for_nha_san_xuat_2G(self,cell_value):
        bts_type = BTSType.objects.get(Name = '2G')
        return super(Excel_3G, self).value_for_Cabinet(self,cell_value,name_ThietBi_attr= 'nha_san_xuat_2G',bts_type = bts_type)
D4_DATE_ONLY_FORMAT_gachngang = '%d-%m-%Y'    
class Excel_to_2g (Excel_2_3g):
    import_ghi_chu_text = 'Ericsson_Database_Ver_160'
    fields_allow_empty_use_function = ['Site_ID_2G_Number','Site_type','import_ghi_chu']
    backwards_sequence =['Site_ID_2G','Site_ID_2G_Number','quan_huyen']
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Database 2G'
    mapping_function_to_value_dict = {#'Ngay_Phat_Song_2G':'value_for_dateField'
                                      }
    manual_mapping_dict = {'Ngay_Phat_Song_2G': u'Phát sóng','Site_Name_1':u'Tên BTS','dia_chi_2G':u'Địa chỉ', 'BSC_2G':u'Tên BSC',\
                    'LAC_2G':u'LAC', 'Nha_Tram':u'Nhà trạm', 'Ma_Tram_DHTT':u'Mã trạm ĐHTT', 'Cell_ID_2G':u'CellId', \
                    'cau_hinh_2G':u'Cấu hình', 'nha_san_xuat_2G':u'Nhà SX', 'Site_ID_2G':u'Tên BTS',\
                    'Long_2G':u'Tọa độ - Kinh độ','Lat_2G':u'Tọa độ - Vĩ độ','quan_huyen':u'Quận/Huyện','tinh':u'Mã tỉnh'}
    def value_for_quan_huyen(self,value):
        if value:
            try:
                instance = QuanHuyen.objects.get(Name = value,tinh = self.obj.tinh )
            except:
                instance = QuanHuyen(Name = value,tinh = self.obj.tinh)
                instance.save()
            return instance
        else:
            return None
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_to_2g,self).value_for_import_ghi_chu(value,'2G: ',insert_index=1)
        return return_value
    def value_for_Ngay_Phat_Song_2G(self,value):
        #print '@@@@@@@@@@@@ ngay phat song 2g',value,type(value)
        rs = datetime.datetime.strptime(value, D4_DATE_ONLY_FORMAT_gachngang)
        #print rs
        return rs
    def value_for_Site_ID_2G_Number(self,value):
        value = self.obj.Cell_ID_2G
        if value:
            value = int(value)
            if value < 1000:
                value = str(value)[-2:]
            else:
                value = str(value)[-3:-1]
            return value
        else:
            return None
    def value_for_BSC_2G(self,value):
        if value:
            try:
                instance = Tram.objects.get(Site_Name_1 = value)
            except Tram.DoesNotExist:
                if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance = Tram(Site_Name_1 = value,Site_type = SiteType.objects.get(Name = u'Site 0 (RNC,BSC)'),ngay_gio_tao = timezone.now(),nguoi_tao = User.objects.get(username = u'rnoc2'))
                instance.save()
        
        if value:
            #print '@@@@@@@@value BSC_2G',value
            try:
                instance = BSCRNC.objects.get(Name = value)
            except BSCRNC.DoesNotExist:
                instance = BSCRNC(Name = value, ngay_gio_tao = timezone.now(),nguoi_tao = User.objects.get(username = 'rnoc2'))
                self.added_foreinkey_types +=1
                if self.added_foreinkey_types  > self.max_length_added_foreinkey_types:
                    raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
                instance.save()
            return instance
        else:
            return None
    def value_for_Site_Name_1 (self,cell_value):
        value = cell_value.replace("2G_","")
        return value
    def value_for_Site_ID_2G(self,cell_value):
        self.obj.active_2G = True
        if self.created_or_update ==1:
            self.obj.nguoi_tao = User.objects.get(username='rnoc2')
            self.obj.ngay_gio_tao = timezone.now()
        if cell_value.startswith('2G_'):
            return None  # return none for not save to database this field
        else:
            cell_value = self.obj.nha_san_xuat_2G.Name[0:3].upper() + '_2G_' + cell_value
            return  cell_value
    def value_for_nha_san_xuat_2G(self,cell_value):
        bts_type = BTSType.objects.get(Name = '2G')
        return_value = super(Excel_to_2g, self).value_for_Cabinet(cell_value,name_ThietBi_attr= 'nha_san_xuat_2G',bts_type = bts_type)
        return  return_value    
    '''
    def value_for_nha_san_xuat_2G(self,cell_value):
        thietbi = ThietBi.objects.get_or_create(Name=cell_value)[0]
        self.added_foreinkey_types.add(thietbi)#set().add
        l = len(self.added_foreinkey_types)
        #print "cabin**",l
        if l >self.max_length_added_foreinkey_types:
            raise ValueError("so luong m2m field qua nhieu, kha nang la ban da chon thu tu field tuong ung voi excel column bi sai")
        self.obj.nha_san_xuat_2G=thietbi
        #self.obj.save()
        return None
    '''
class Excel_to_3g_location (Excel_2_3g):
    fields_allow_empty_use_function = ['Site_type']
    allow_create_one_instance_if_not_exit = False
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_ID_3G'
    worksheet_name = u'3G Site Location'
    mapping_function_to_value_dict ={}
    manual_mapping_dict = {'Site_ID_3G':u'Site ID','dia_chi_3G':u'Location','Long_3G':u'Long','Lat_3G':u'Lat',}
    def value_for_Site_ID_3G(self,cell_value):
        value = 'ERI_3G_' + cell_value
        return value
class Excel_to_2g_config_SRAN (Excel_2_3g):
    fields_allow_empty_use_function = ['Site_type']
    begin_row=39
    auto_map = False
    just_create_map_field = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'2G SRAN HCM Config'
    mapping_function_to_value_dict ={}
    manual_mapping_dict = {'Site_Name_1':u'RSITE','TG_Text':u'TG','TRX_DEF':u'TRX DEF'}
    def value_for_TG_Text(self,value):
        self.obj.active_2G = True
        if self.created_or_update ==1:
            self.obj.nguoi_tao = User.objects.get(username='rnoc2')
            self.obj.ngay_gio_tao = timezone.now()
        rs = re.findall('RXOTG-(\d+)',value)
        if len(rs)==1:
            self.obj.TG = rs[0]
        elif len(rs)>1:
            self.obj.TG = rs[0]
            self.obj.TG_1800 = rs[1]
                
        return value
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('2G_','')
        return cell_value
class Excel_NSM(Excel_2_3g):
    fields_allow_empty_use_function = ['Site_type','import_ghi_chu']
    begin_row=1
    just_create_map_field = False
    auto_map = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'NSN Database'
    mapping_function_to_value_dict ={'Ngay_Phat_Song_3G':'value_for_common_datefield','IUB_VLAN_ID':'value_for_common_VLAN_ID'}
    manual_mapping_dict = {'Site_Name_1':u'3G Site Name','Site_ID_3G':u'3G Site Name','Cabinet':u'Type',\
                    'Ngay_Phat_Song_3G':u'Ngày PS U900','RNC':u'RNC name','IUB_VLAN_ID':u'VLAN ID','IUB_DEFAULT_ROUTER':u'GW IP ',\
                    'IUB_HOST_IP':u'IP','MUB_SUBNET_PREFIX':u'Network IP','MUB_DEFAULT_ROUTER':u'TRS IP',\
                    'ntpServerIpAddressPrimary':u'NTP Primary IP','ntpServerIpAddressSecondary':u'NTP Secondary  IP','tinh':u'Province'
                    }
    def value_for_Cabinet(self,cell_value):
        cell_value = 'NSM'
        bts_type = BTSType.objects.get(Name = '3G')
        return_value = super(Excel_NSM, self).value_for_Cabinet(cell_value,bts_type=bts_type)
        return  return_value
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('3G_','')
        return cell_value
    def value_for_Site_ID_3G(self,cell_value):
        self.obj.active_3G = True
        if self.created_or_update ==1:
            self.obj.nguoi_tao = User.objects.get(username='rnoc2')
            self.obj.ngay_gio_tao = timezone.now()
        #cell_value = cell_value.replace('3G_','NSM_')
        cell_value = 'NSM_'+ cell_value
        return cell_value
    
class Excel_ALU(Excel_2_3g):
    fields_allow_empty_use_function = ['nguoi_tao','ngay_gio_tao','Site_type']
    begin_row=3
    just_create_map_field = False
    auto_map = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Database_ALU'
    mapping_function_to_value_dict ={'Ngay_Phat_Song_3G':'value_for_common_datefield','IUB_VLAN_ID':'value_for_common_VLAN_ID','MUB_VLAN_ID':'value_for_common_VLAN_ID'}
    manual_mapping_dict = {'Site_Name_1':u'Tên trạm (ALU)','Site_ID_3G':u'Tên trạm (ALU)','Cabinet':u'RNC',\
                    'RNC':u'RNC',
                    'IUB_VLAN_ID':u'Iub Vlan','IUB_SUBNET_PREFIX':u'Iub Subnet',
                    'IUB_DEFAULT_ROUTER':u'Iub Default Router','IUB_HOST_IP':u'Iub Host',
                    'MUB_VLAN_ID':u'Mub Vlan','MUB_SUBNET_PREFIX':u'Mub Subnet',\
                    'MUB_HOST_IP':u'Mub Host','MUB_DEFAULT_ROUTER':u'Mub Default Router',\
                   
                    }
    # bat dau dem column tu 0
    manual_mapping_dict = {'Site_Name_1':2,'Site_ID_3G':2,'Cabinet':8,\
                    'RNC':8,\
                    'IUB_VLAN_ID':11,'IUB_SUBNET_PREFIX':12,\
                    'IUB_DEFAULT_ROUTER':13,'IUB_HOST_IP':14,\
                    'MUB_VLAN_ID':15,'MUB_SUBNET_PREFIX':16,\
                    'MUB_HOST_IP':18,'MUB_DEFAULT_ROUTER':17,\
                   
                    }
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('3G_','')
        return cell_value
    def value_for_Site_ID_3G(self,cell_value):
        self.obj.active_3G = True
        
        cell_value = 'ALU_'+ cell_value
        return cell_value
    def value_for_Cabinet(self,cell_value):
        cell_value = 'Alcatel'
        bts_type = BTSType.objects.get(Name = '3G')
        return_value = super(Excel_ALU, self).value_for_Cabinet(cell_value,bts_type=bts_type)
        return  return_value
class Excel_ALU_tuan(Excel_2_3g):
    fields_allow_empty_use_function = ['nguoi_tao','ngay_gio_tao','Cabinet','Site_type','import_ghi_chu']
    begin_row=0
    just_create_map_field = False
    auto_map = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Sheet1'
    mapping_function_to_value_dict ={'Ngay_Phat_Song_3G':'value_for_common_datefield','IUB_VLAN_ID':'value_for_common_VLAN_ID','MUB_VLAN_ID':'value_for_common_VLAN_ID'}
    manual_mapping_dict = {'Site_Name_1':u'NodeB Name','Site_ID_3G':u'NodeB Name',\
                    'RNC':u'RNC Nam','IUB_DEFAULT_ROUTER':u'Iub Default GW NodeB','IUB_HOST_IP':u'Iub NodeB Ip add',
                    'IUB_VLAN_ID':u'Iub Vlan','IUB_SUBNET_PREFIX':u'Iub Default GW NodeB',
                    'MUB_VLAN_ID':u'Mub Vlan','MUB_SUBNET_PREFIX':u'Mub Default GW NodeB',\
                    'MUB_HOST_IP':u'Mub NodeB Ip add','MUB_DEFAULT_ROUTER':u'Mub Default GW NodeB',\
                    'tinh':'Tinh',
                   
                    }
    
    def value_for_Site_Name_1 (self,cell_value):
        cell_value = cell_value.replace('3G_','')
        return cell_value
    def value_for_Site_ID_3G(self,cell_value):
        self.obj.active_3G = True
        cell_value = 'ALU_'+ cell_value
        return cell_value
    def value_for_Cabinet(self,cell_value):
        cell_value = 'ALU'
        bts_type = BTSType.objects.get(Name = '3G')
        return_value = super(Excel_ALU_tuan, self).value_for_Cabinet(cell_value,bts_type = bts_type)
        return  return_value

class Excel_4G(Excel_2_3g):
    fields_allow_empty_use_function = ['Site_type','import_ghi_chu']
    #begin_row=3
    just_create_map_field = False
    auto_map = False
    update_or_create_main_item = 'Site_Name_1'
    worksheet_name = u'Ericsson 4G'
    mapping_function_to_value_dict ={'eNodeB_ID_DEC':'value_for_common_VLAN_ID'}
    manual_mapping_dict = {'eNodeB_Name':u'eNodeB_Name','Site_Name_1':u'eNodeB_Name','eNodeB_ID_DEC':u'eNodeB_ ID(DEC)','eNodeB_Type':u'eNodeB_Type',     }
    def value_for_tinh(self, value):
        value = self.obj.Site_Name_1[-3:]
        return super(Excel_4G,self).value_for_tinh(value)
    def value_for_import_ghi_chu(self,value):
        return_value = super(Excel_4G,self).value_for_import_ghi_chu(value,'4G: ',insert_index=3)
        return return_value
    def value_for_Site_Name_1 (self,cell_value):
        #results = re.findall('4G_(.*?_', cell_value)
        cell_value = cell_value.replace('4G_','')
        return cell_value
    def value_for_eNodeB_Type(self,cell_value):
        self.obj.active_4G = True
        if self.created_or_update ==1:
            self.obj.nguoi_tao = User.objects.get(username='rnoc2')
            self.obj.ngay_gio_tao = timezone.now()
        bts_type = BTSType.objects.get(Name = '4G')
        return_value = super(Excel_4G, self).value_for_Cabinet(cell_value,name_ThietBi_attr = 'eNodeB_Type',bts_type = bts_type)
        return  return_value    
from django.contrib.auth.models import User, Group
from django.contrib.auth.models import Permission
from django.contrib.contenttypes.models import ContentType
def create_user():
    workbook = xlrd.open_workbook(MEDIA_ROOT+ '/document/DanhSachEmail.xls')
    worksheet = workbook.sheet_by_name(u'Sheet3')
    num_rows = worksheet.nrows - 1
    row_index = -1
    while row_index < num_rows:
        row_index += 1
        username =   read_excel_cell(worksheet, row_index, 6)
        sdt  =   read_excel_cell(worksheet, row_index, 5)
        groupname =   read_excel_cell(worksheet, row_index, 7)
        ex = User.objects.get_or_create (username = username)
        user = ex[0]
        if ex[1]:#Neu user = New                                
            user.set_password(username)
            user.save()                          
        group = Group.objects.get_or_create (name = groupname)[0]
        group.user_set.add(user)
        ex = UserProfile.objects.get_or_create(user =user)
        if ex[1]:
            profile = ex[0]
            profile.so_dien_thoai=sdt
            profile.color_code = "#%06x" % random.randint(0, 0xFFFFFF)
            ca_truc = CaTruc.objects.latest('id')
            profile.ca_truc = ca_truc
            profile.save()
    more_users = ['ductu','rnoc2']
    for username in more_users:
        ex = User.objects.get_or_create (username = username)
        user = ex[0]
        if ex[1]:#Neu user = New                                
            user.set_password(username)
            user.save()  
        ex = UserProfile.objects.get_or_create(user =user)
        if ex[1]:
            profile = ex[0]
            profile.color_code = "#%06x" % random.randint(0, 0xFFFFFF)
            ca_truc = CaTruc.objects.latest('id')
            profile.ca_truc = ca_truc
            profile.save()
 

def grant_permission_to_group():
    content_type = ContentType.objects.get_for_model(Mll)
    name_and_codes = [('d4_create_truc_ca_permission','Can truc ca'),('can add on modal code','can add on modal')]
    truc_ca_group = Group.objects.get_or_create (name = 'truc_ca')[0]
    for x in name_and_codes:
        permission = Permission.objects.get_or_create(codename=x[0],
                                           name=x[1],
                                           content_type=content_type)[0]
    
        truc_ca_group.permissions.add(permission)
def grant_permission_admin():
    content_type = ContentType.objects.get_for_model(Mll)
    code_and_names= [('d4_admin','Can admin')]
    user = User.objects.get (
                                            username = 'rnoc2',
                                            )
    for x in code_and_names:
        permission = Permission.objects.get_or_create(codename=x[0],
                                           name=x[1],
                                           content_type=content_type)[0]
    
        user.user_permissions.add(permission)
        
def check_permission_of_group():
    for username in ['rnoc2','lucvk']:
        user = User.objects.get_or_create (
                                            username = username,
                                            )[0]
        permission = Permission.objects.get(codename='d4_create_truc_ca_permission')
        #print 'username,user.has_perm',username,user.has_perm('drivingtest.d4_create_truc_ca_permission')


            
              
from django.template import Context,Template 

def tao_script(instance_site,ntpServerIpAddressPrimary = '',ntpServerIpAddressSecondary = '',\
                         ntpServerIpAddress1="",ntpServerIpAddress2=""):
    if (ntpServerIpAddressPrimary=='' or ntpServerIpAddress1==""):
        return None
    #print 'hello, wellcome to download'
    Cabinet = instance_site.Cabinet
    is_luu_o_cung_moi_file_output_rieng = True
    save_type = 'temporary_achive_output_script'#or save_type = 'disk_achive_output_script',khong xai, 
    #chi de hieu rang achive object co the ghi len o cung hoac len file tam
    now = datetime.datetime.now()
    Site_ID_3G= instance_site.Site_ID_3G
    instance_site.now = now
    return_file_lists = []
    achive_path=None
    instance_site.ntpServerIpAddressPrimary = ntpServerIpAddressPrimary
    instance_site.ntpServerIpAddressSecondary = ntpServerIpAddressSecondary
    instance_site.ntpServerIpAddress1 = ntpServerIpAddress1
    instance_site.ntpServerIpAddress2 = ntpServerIpAddress2
    template_files =[]
    if "RBS6" in Cabinet.Name:
        type_rbs = "6000"
        path_template_directory = MEDIA_ROOT+ '/document/template_script/6000/'
    elif "RBS3" in Cabinet.Name:
        path_template_directory = MEDIA_ROOT+ '/document/template_script/3000/'
        type_rbs = "3000"
    for root, dirs, files in os.walk(path_template_directory):
        for file in files:
            template_files.append(file)
    for counts,template_file in enumerate(template_files):
        path_to_1_template_file =  path_template_directory + template_file
        template = read_file_from_disk (path_to_1_template_file)
        t = Template(template)
        c = Context({'site3g':instance_site})
        output = t.render(c)
        fname = Site_ID_3G + '_' + template_file
        folder_name = '5484692'
        new_directory_path = MEDIA_ROOT+ '/for_user_download_folder/' + folder_name + '/'
        if save_type == 'save_to_disk_3_file':# chi luu o cung trong giao dien console nay
            if not os.path.exists(new_directory_path): os.makedirs(new_directory_path)
            filepath = new_directory_path  + fname
            return_file_lists.append(folder_name + '/' +  fname)
            save_file_to_disk(filepath,output,1)
        else:
            if counts==0:
                if save_type =='disk_achive_output_script':
                    achive_path = new_directory_path + Site_ID_3G +'.zip'#achive_path den o cung
                elif  save_type == 'temporary_achive_output_script':# dang dung
                    achive_path = tempfile.TemporaryFile() # this time achive_path is template object file,achive_path la 1 object template
                archive_object = zipfile.ZipFile(achive_path, 'w', zipfile.ZIP_DEFLATED)# tao ra 1 object achive de write len path
            if is_luu_o_cung_moi_file_output_rieng:#luu o cung 3 file rieng re, de doi chieu voi download o giao dien web
                if not os.path.exists(new_directory_path):
                    os.makedirs(new_directory_path)
                filepath = new_directory_path  + fname
                save_file_to_disk(filepath,output,1)
            archive_object.writestr(fname, output)#write to object theo kieu data voi file name la fname
    return return_file_lists, achive_path, type_rbs # achive_path become tempt zip file

    #grant_permission_admin()

def create_ca_truc():
    for ca_truc_name in ['Moto','Alu','Huawei','Sran']:
        ex = CaTruc.objects.get_or_create(Name=ca_truc_name)
        instance = ex[0]
        if ex[1]:
            instance.is_duoc_tao_truoc = True
            #instance.nguoi_tao = User.objects.get(username = "rnoc2")
            #instance.ngay_tao = timezone.now()
            instance.save()
def create_type_site():
    for x in ['Site 0 (RNC,BSC)',u'Site thường']:
        try:
            instance = SiteType.objects.get(Name = x)
        except SiteType.DoesNotExist:
            instance = SiteType(Name=x)
            instance.save()
def create_type_bts():
    for x in [u'2G',u'3G',u'4G',u'ALL Band']:
        try:
            instance = BTSType.objects.get(Name = x)
        except BTSType.DoesNotExist:
            instance = BTSType(Name=x)
            instance.save()
                        

def delete_edithistory_table3g():
    EditHistory.objects.filter(modal_name='Tram').delete()
from openpyxl import load_workbook
def export_excel_bcn(querysets = None,yesterday_or_other = None):
    if yesterday_or_other:
        if yesterday_or_other == 'Today':
            min_select_day = datetime.date.today() #datetime.datetime.now9).date()
            querysets =querysets.filter(gio_mat__month=min_select_day.month,gio_mat__year=min_select_day.year,gio_mat__day = min_select_day.day)
        elif yesterday_or_other == 'Yesterday':
            min_select_day = datetime.date.today() - datetime.timedelta(days = 1)
            querysets =querysets.filter(gio_mat__month=min_select_day.month, gio_mat__year=min_select_day.year,gio_mat__day = min_select_day.day)
    else:#theotable:
        min_select_day = awaredate_time_to_local(querysets.aggregate(Min('gio_mat'))['gio_mat__min']).date()
    response = HttpResponse()
    path_or_file_or_response = response
    response['Content-Disposition'] = 'attachment; filename="bcn.xls"'
    wb = load_workbook(MEDIA_ROOT  + '/document/BCN_MLL_2G_3G_4G_02-05-2016.xlsx')
    SHEETS = ["BCN_2G","BCN_3G"]
    for sheet in SHEETS:
        ws = wb[sheet]
        max_row = ws.max_row
        if sheet == "BCN_2G":
            begin_row = 47
            range_xls  = 'A%s:J%s'%(str(begin_row),str(max_row))
            name_for_filter = '2G'
        elif sheet == "BCN_3G" :
            begin_row = 46
            range_xls  = 'A%s:I%s'%(str(begin_row),str(max_row))
            name_for_filter = '3G'
        for row in ws.iter_rows(range_xls):
            for cell in row:
                cell.value = None
        bcns_2g = querysets.filter(BTS_Type__Name= name_for_filter).order_by('gio_mat').exclude(code_loi = 8).exclude(code_loi = 7)
        #list(chain(kq_searchs, kq_searchs_one_contain))
        #bcns_2g = querysets.filter(BTS_Type__Name= name_for_filter).exclude(code_loi = 8).exclude(code_loi = 7).order_by('gio_mat')
        previous_day =   min_select_day - datetime.timedelta(days = 1)
        previousday_max_time = local_a_naitive(datetime.datetime.combine(previous_day,datetime.time(23,59,59)))
        group = Q(BTS_Type__Name= name_for_filter,gio_mat__month=previous_day.month,gio_mat__year=previous_day.year,gio_mat__day = previous_day.day, gio_tot__gte = previousday_max_time )
        bcns_2g_homtruocs =  BCNOSS.objects.filter(group).exclude(code_loi = 8).exclude(code_loi = 7)
        bcns_2g = list(chain(bcns_2g, bcns_2g_homtruocs))
        for count,bcn_record_row in enumerate(bcns_2g):
            ws.cell(row = begin_row + count, column = 1).value = 2
            ws.cell(row = begin_row + count, column = 3).value = bcn_record_row.object
            ws.cell(row = begin_row + count, column = 4).value = bcn_record_row.BSC_or_RNC.Name
            report_row_day = awaredate_time_to_local(bcn_record_row.gio_mat).date()
            if report_row_day == previous_day:
                report_row_day = min_select_day
                gio_mat = local_a_naitive(datetime.datetime.combine(report_row_day,datetime.time(0,0,0)))
            else:
                gio_mat = bcn_record_row.gio_mat
            ws.cell(row = begin_row + count, column = 5).value = awaredate_time_to_local(gio_mat).time().strftime(TIME_FORMAT_FOR_BCN)
            ws.cell(row = begin_row + count, column = 2).value = report_row_day.strftime(DATE_FORMAT_FOR_BCN)
            row_max_time = local_a_naitive(datetime.datetime.combine(report_row_day,datetime.time(23,59,59)))
            if bcn_record_row.gio_tot:
                gio_tot = awaredate_time_to_local(bcn_record_row.gio_tot)
                if gio_tot > row_max_time:
                    gio_tot = row_max_time
            else:#row_max_time neu khong co gio tot
                gio_tot = awaredate_time_to_local(row_max_time)
            ws.cell(row = begin_row + count, column = 6).value = gio_tot.strftime(TIME_FORMAT_FOR_BCN)
            tong_thoi_gian = int(round((gio_tot -gio_mat).seconds/60.0))
            ws.cell(row = begin_row + count, column = 7).value = tong_thoi_gian
            ws.cell(row = begin_row + count, column = 8).value = str(bcn_record_row.code_loi)
            ws.cell(row = begin_row + count, column = 9).value = bcn_record_row.vnp_comment 
            if sheet == "BCN_2G":
                ws.cell(row = begin_row + count, column = 10).value = bcn_record_row.gio_canh_bao_ac.strftime(DATETIME_FORMAT_FOR_BCN) if bcn_record_row.gio_canh_bao_ac else None
                ws.cell(row = begin_row + count, column = 11).value = bcn_record_row.object[-3:] 
            else:
                ws.cell(row = begin_row + count, column = 10).value = bcn_record_row.object[-3:] 
    
    
    wb.save(path_or_file_or_response)
    return path_or_file_or_response
    
def init_rnoc():
    create_ca_truc()#1
    create_user()#2
    import_database_4_cai_new(['ExcelImportTrangThai'])#3
    import_database_4_cai_new(['ExcelImportSuCo'])#4
    import_database_4_cai_new(['ExcelImportDuAn'])#5
    import_database_4_cai_new(['ExcelImportNguyenNhan'])#6
    #import_database_4_cai_new(['ExcelImportThietBi'])#7
    import_database_4_cai_new(['ExcelImportFaultLibrary'])#8
    import_database_4_cai_new(['ExcelImportThaoTacLienQuan'])#9
    import_database_4_cai_new(['ExcelImportLenh'])#10
    import_database_4_cai_new(['ExcelImportDoiTac'])#11
    import_database_4_cai_new(['ExcelImportDoiTac_ungcuu'] )
    import_database_4_cai_new(['ImportTinh'] )
    import_database_4_cai_new(['ImportTinh_diaban'] )
    create_type_site()
    create_type_bts()
    import_database_4_cai_new(['Import_RNC_Tram'] )
from django.db.models.aggregates import Sum, Min
from django.db.models import Avg
from dateutil import rrule
def thong_ke_theo_ma_loi(qs,code_loi,tong_thoi_gian_mat,so_lan_mat_lien_lac):
        qs_loi = qs.filter(code_loi=code_loi)
        so_lan_mat_lien_lac_code = qs_loi.count()
        if so_lan_mat_lien_lac_code==0:
            so_lan_mat_lien_lac_percent = 0
            mat_dien_sum = 0
            avg_mat_dien = '_'
        else:
            so_lan_mat_lien_lac_percent = str(round(so_lan_mat_lien_lac_code*100/float(so_lan_mat_lien_lac),2))
            mat_dien_sum = qs_loi.aggregate(Sum('tong_thoi_gian'))['tong_thoi_gian__sum']
            avg_mat_dien =round(qs_loi.aggregate(Avg('tong_thoi_gian'))['tong_thoi_gian__avg'],2)
        ket_luan = mark_safe(u'{3} lần<span style="color:red">({4}%)</span>|{0} Phút({1}%)|{2} phút/lần'.format(mat_dien_sum,str(round(mat_dien_sum*100/float(tong_thoi_gian_mat),2)),avg_mat_dien,
                                                                         so_lan_mat_lien_lac_code,so_lan_mat_lien_lac_percent))
        return ket_luan

class ApiDataset(object):#thong ke bao cao ngay
    
    def __init__(self,BTS_type = '3G',MONTHLY_or_DAILY= 'DAILY',bg=None,end=None):
        if MONTHLY_or_DAILY== 'MONTHLY':
            bg = datetime.datetime(2016, 1, 1, 15, 29, 43, 79060)
            end = datetime.datetime(2016, 5, 12, 15, 29, 43, 79060)
        elif MONTHLY_or_DAILY== 'DAILY':
            end = datetime.datetime.now()
            bg  = end -datetime.timedelta(days=7)
        self.ls = rrule.rrule(getattr(rrule,MONTHLY_or_DAILY), dtstart=bg, until=end)
        self.BTS_type = BTS_type
        self.MONTHLY_or_DAILY = MONTHLY_or_DAILY
        print 'self.lsself.lsself.lsself.lsself.lslen(self.ls)'
    '''
    def cache_data(self):
        # Access API and cache returned data on object.
        if self.data is None:
            self.data = 1
    '''
    def __iter__(self):
        so_lan_mat_lien_lac_prev = None
        for x in self.ls:
            if self.MONTHLY_or_DAILY== 'MONTHLY':
                thang_nam = x.strftime('%m/%Y')
            elif self.MONTHLY_or_DAILY== 'DAILY':
                thang_nam = x.strftime('%d/%m/%Y')
            karg = {'gio_mat__month':x.month,'gio_mat__year':x.year,'BTS_Type__Name':self.BTS_type}
            if self.MONTHLY_or_DAILY== 'DAILY':
                karg.update({'gio_mat__day':x.day})
            qs = BCNOSS.objects.filter(**karg)
            so_lan_mat_lien_lac = qs.count()
            if not so_lan_mat_lien_lac_prev:
                so_lan_mat_lien_lac_Increase_or_descrease = u'_'
            else:
                so_lan_mat_lien_lac_Increase_or_descrease = ((so_lan_mat_lien_lac - so_lan_mat_lien_lac_prev )/float(so_lan_mat_lien_lac_prev))*100
            if isinstance(so_lan_mat_lien_lac_Increase_or_descrease, float):
                if so_lan_mat_lien_lac_Increase_or_descrease > 0 :
                    so_lan_mat_lien_lac_Increase_or_descrease = u'{0}%'.format(u'<span class="glyphicon glyphicon-arrow-up" style="color:red"></span>+%.1f'%so_lan_mat_lien_lac_Increase_or_descrease)
                else:
                    so_lan_mat_lien_lac_Increase_or_descrease = u'{0}%'.format(u'%.1f'%so_lan_mat_lien_lac_Increase_or_descrease)
            so_lan_mat_lien_lac_prev = so_lan_mat_lien_lac
            tong_thoi_gian_mat = qs.aggregate(Sum('tong_thoi_gian'))['tong_thoi_gian__sum']
            if so_lan_mat_lien_lac:
                thoi_gian_trung_binh_1_lan_mat ="{0:.2f}".format( round(qs.aggregate(Avg('tong_thoi_gian'))['tong_thoi_gian__avg'],1))
            else:
                thoi_gian_trung_binh_1_lan_mat = 0
                
            if self.BTS_type == '2G':
                tong_so_luong_tram_2g = Tram.objects.filter(active_2G=True).count()
            else:
                tong_so_luong_tram_2g = Tram.objects.filter(active_3G=True).count()
            try :
                thoi_luong_mat_trung_binh_cua_1_tram_trong_thang = u"%.1f"%(tong_thoi_gian_mat/float(tong_so_luong_tram_2g))
                #thoi_luong_mat_trung_binh_cua_1_tram_trong_thang = tong_so_luong_tram_2g
            except ZeroDivisionError:
                thoi_luong_mat_trung_binh_cua_1_tram_trong_thang =u'_'
            except TypeError:
                thoi_luong_mat_trung_binh_cua_1_tram_trong_thang =u'_'

            if so_lan_mat_lien_lac==0:
                mat_dien_tong = u'_'
                truyen_dan_tinh_tong = u'_'
                thiet_bi_tong = u'_'
            else:
                mat_dien_tong = thong_ke_theo_ma_loi(qs,1,tong_thoi_gian_mat,so_lan_mat_lien_lac)
                truyen_dan_tinh_tong = thong_ke_theo_ma_loi(qs,5,tong_thoi_gian_mat,so_lan_mat_lien_lac)
                thiet_bi_tong = thong_ke_theo_ma_loi(qs,3,tong_thoi_gian_mat,so_lan_mat_lien_lac)
            so_lan_mat_lien_lac_txt = u'{0} Lần {1}'.format(so_lan_mat_lien_lac,so_lan_mat_lien_lac_Increase_or_descrease)
            data_item = {'thang_nam':thang_nam,'so_lan_mat_lien_lac':so_lan_mat_lien_lac_txt,\
                         'tong_thoi_gian_mat':u'%s phút'%tong_thoi_gian_mat,'thoi_gian_trung_binh_1_lan_mat':u'%s phút'%thoi_gian_trung_binh_1_lan_mat,\
                         'thoi_luong_mat_trung_binh_cua_1_tram_trong_thang':u'%s phút(tổng số trạm %s:%s)'%(thoi_luong_mat_trung_binh_cua_1_tram_trong_thang,self.BTS_type,tong_so_luong_tram_2g),\
                         'mat_dien_tong':mat_dien_tong,'truyen_dan_tinh_tong':truyen_dan_tinh_tong,'thiet_bi_tong':thiet_bi_tong,
                         }
            yield data_item 

    def __len__(self):
        return len(self.ls)


def thongkebcn_generator():
    
    bg = datetime.datetime(2014, 1, 1, 15, 29, 43, 79060)
    end = datetime.datetime(2016, 5, 12, 15, 29, 43, 79060)
    ls = rrule.rrule(rrule.MONTHLY, dtstart=bg, until=end)
    data_item = {}
    for x in ls:
        thang_nam = x.strftime('%m/%Y')
        qs = BCNOSS.objects.filter(gio_mat__month=x.month,gio_mat__year=x.year).exclude(code_loi = 8)
        so_lan_mat_lien_lac = qs.count()
        tong_thoi_gian_mat = qs.aggregate(Sum('tong_thoi_gian'))['tong_thoi_gian__sum']
        try:
            thoi_gian_trung_binh_1_lan_mat ="{0:.2f}".format( round(qs.aggregate(Avg('tong_thoi_gian'))['tong_thoi_gian__avg'],2))
        except TypeError:
            thoi_gian_trung_binh_1_lan_mat = None
        tong_so_luong_tram_2g = Tram.objects.filter(active_2G=True).count()
        try :
            thoi_luong_mat_trung_binh_cua_1_tram_trong_thang = u"%.2f"%(tong_thoi_gian_mat/float(tong_so_luong_tram_2g))
            #thoi_luong_mat_trung_binh_cua_1_tram_trong_thang = tong_so_luong_tram_2g
        except ZeroDivisionError:
            thoi_luong_mat_trung_binh_cua_1_tram_trong_thang =None
        except TypeError:
            thoi_luong_mat_trung_binh_cua_1_tram_trong_thang =None
        data_item = {'thang_nam':thang_nam,'so_lan_mat_lien_lac':so_lan_mat_lien_lac,\
                     'tong_thoi_gian_mat':tong_thoi_gian_mat,'thoi_gian_trung_binh_1_lan_mat':thoi_gian_trung_binh_1_lan_mat,'thoi_luong_mat_trung_binh_cua_1_tram_trong_thang':thoi_luong_mat_trung_binh_cua_1_tram_trong_thang}
        yield data_item    
        
def import_database_4_cai_new (runlists,workbook = None,import_ghi_chu = None):
        #'Excel_3G','Excel_to_2g',
        DB3G_SHEETS = ['Excel_3G','Excel_to_2g','Excel_to_2g_config_SRAN','Excel_to_3g_location','Excel_4G']
        is_db3g = False
        if 'ALL' in runlists:
            runlists.remove('ALL')
            runlists.extend(DB3G_SHEETS)
            runlists = unique_list(runlists)
        runlists_copy = runlists[:]
        runlists_reorder = []
        for x in DB3G_SHEETS:
            if x in runlists:
                is_db3g = True
                runlists_reorder.append(x)
                runlists_copy.remove(x)
        if is_db3g:
            for x in runlists_copy:
                runlists_reorder.append(x)
            runlists = runlists_reorder
            is_already_read_db3g_file  = False
        if workbook:
            for class_func_name in runlists:
                running_class = eval(class_func_name)
                return running_class(workbook = workbook,import_ghi_chu=import_ghi_chu).thong_bao
        else:
            for class_func_name in runlists:
                if class_func_name in DB3G_SHEETS:
                    if not is_already_read_db3g_file:
                        path = MEDIA_ROOT+ u'/document/Ericsson_Database_Ver_161.xlsx'
                elif class_func_name =='ImportRNC':
                    path = MEDIA_ROOT+ '/document/rnc.xls'
                elif class_func_name =='Excel_NSM':
                    path = MEDIA_ROOT+ '/document/NSN_Database_version_4.xlsx'
                elif class_func_name =='Excel_ALU':
                    path = MEDIA_ROOT+ '/document/Database_ALU lot 1-2 -3 den NGAY  5-8-2015 .xls'
                elif class_func_name =='Excel_ALU_tuan':
                    path = MEDIA_ROOT+ '/document/alu_tuan.xlsx'
                elif class_func_name =='ExcelImportDoiTac_ungcuu':
                    path = MEDIA_ROOT+ '/document/danh sach nv xuong quan ly dia ban.xls'
                elif class_func_name =='ImportTinh':
                    path = MEDIA_ROOT+ '/document/danh sach nv xuong quan ly dia ban.xls'
                elif class_func_name =='ImportTinh_diaban':
                    path = MEDIA_ROOT+ '/document/To Ung cuu_New.tu.xls'
                elif class_func_name =='Import_RNC_Tram':
                    path = MEDIA_ROOT+ '/document/Table_BSCRNC.xls'
                elif class_func_name =='Excel_3G':
                    path = MEDIA_ROOT+ '/document/3g_test.xls'
                elif class_func_name =='Excel_to_2g':
                    path = MEDIA_ROOT+ '/document/2g_test.xls'
                elif class_func_name =='ImportBCN2G':
                    path  = '/media/ductu/2CB8955EB895277C/tu0101den1205.xls'
                elif class_func_name =='ImportBCN2G_SRAN':
                    path = '/home/ductu/Documents/Downloads/2sran.xls'
                elif class_func_name =='ImportBCN3G':
                    path = '/home/ductu/Documents/Downloads/3g1.xls'
                elif class_func_name =='ImportBCN3G_ALU':
                    path = '/home/ductu/Documents/Downloads/ALU.xls'
                elif class_func_name =='ImportBCN3G_NSM':
                    path = '/home/ductu/Documents/Downloads/NSM.xls'
                elif class_func_name =='ExcelChung':
                    path = '/home/ductu/Documents/Downloads/Table_Tram.xls'
                else: 
                    rs = re.match('^ExcelImport(.*?)$',class_func_name)
                    try:
                        classname = rs.group(1)
                        path = MEDIA_ROOT+ '/document/Table_%s.xls'%classname
                    except AttributeError:
                        raise ValueError('khong ton tai file name nao nhu the trong thu muc media')
                if class_func_name in DB3G_SHEETS:
                    if not is_already_read_db3g_file:     
                        workbook= xlrd.open_workbook(path)
                        is_already_read_db3g_file = True
                else:
                    workbook= xlrd.open_workbook(path)
                running_class = eval(class_func_name)
                import_ghi_chu = path[path.rfind('/')+1:]
                thong_bao = running_class(workbook = workbook,import_ghi_chu=import_ghi_chu).thong_bao
        return thong_bao
if __name__ == '__main__':
    #ApiDataset()
    #create_user()#2
    #import_database_4_cai_new(['Excel_3G','Excel_to_2g','Excel_4G','Excel_to_2g_config_SRAN','Excel_to_3g_location','ImportRNC','Excel_NSM','Excel_ALU_tuan',] )
    #export_excel_bcn()
    #thongkebcn_generator()
    x = datetime.date.today()
    x = local_a_naitive(datetime.datetime.combine(x,datetime.time(23,59,59)))
    print x
    pass
    
    
    
        
        

    
    