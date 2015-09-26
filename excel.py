'''
from openpyxl import load_workbook
wb = load_workbook(filename = '/home/ductu/Downloads/3g.xlsx',read_only=True)
ws = wb['Database 3G']
d = ws.cell(row = 4, column = 2)
print d.value
count=1
for row in ws.iter_rows('A2:AJ3'):
    print 'number of row ', count
    count+=1
    for cell in row:
        print cell.value
'''
from pyexcelerate import Workbook

wb = Workbook()
ws = wb.new_sheet("test")
ws.range("B2", "C3").value = [[1, 2], [3, 4]]
wb.save("output.xlsx")